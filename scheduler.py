"""
Excel Room Scheduler 

HARD constraints
- Room size: meeting must be in MinRoomSize or larger
- Room must be open
- 15-min grid
- Turnover buffer: require 15 minutes between meetings in the SAME ROOM

SOFT preferences (scored; higher importance = higher penalty)
1) Avoid overlap (with meetings listed in AvoidOverlapWith)  [MOST IMPORTANT]
2) Start / end times (EarliestStart / LatestEnd)            [2nd]
3) Date (PreferredDate)                                     [3rd]
4) Room (PreferredRooms)                                    [LEAST IMPORTANT]

Sheets expected
Rooms:
  RoomName | Size | Date | OpenStart | OpenEnd

Meetings:
  MeetingName | DurationMin | MinRoomSize | Priority | EarliestStart | LatestEnd
  AvoidOverlapWith | PreferredRooms | PreferredDate

Outputs
- Schedule sheet: polished grid with thick room bars, hour separators, merged meeting blocks,
  darker crosshatched closed times.
- Meetings sheet: Status | UnscheduledReason | AssignedRoom | AssignedStart | AssignedEnd | Notes
  - UNSCHEDULED highlighted red (only when no feasible slot exists under HARD constraints)
  - Scheduled but missed preferences highlighted yellow with explicit Notes
  - Scheduled with all preferences met highlighted green

Run:
  python scheduler.py input.xlsx --output output.xlsx
"""

import argparse
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from typing import Dict, List, Optional, Tuple, Set

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------
# Settings
# -----------------------------
SIZE_ORDER = {"small": 1, "med": 2, "large": 3, "xl": 4}
SLOT_MINUTES = 15
ROOM_TURNOVER_MINUTES = 15  # HARD: gap between meetings in same room

# Preference scoring weights (bigger = more important)
W_OVERLAP = 1_000_000   # highest
W_TIME = 100_000        # second
W_DATE = 10_000         # third
W_ROOM = 1_000          # least

# Extra scaling for "how bad" a miss is
W_TIME_PER_MIN = 50     # multiplies minutes outside preferred window
W_OVERLAP_PER_MIN = 500 # multiplies overlap minutes with avoided meetings

# -----------------------------
# Data models
# -----------------------------
@dataclass(frozen=True)
class Room:
    name: str
    size: str


@dataclass
class OpenBlock:
    room_name: str
    day: date
    start: time
    end: time


@dataclass
class Meeting:
    name: str
    duration_min: int
    min_size: str
    priority: int
    preferred_day: Optional[date]          # PreferredDate (SOFT)
    earliest: Optional[time]               # EarliestStart (SOFT)
    latest_end: Optional[time]             # LatestEnd (SOFT)
    avoid_overlap_with: List[str]          # AvoidOverlapWith (SOFT, highest weight)
    preferred_rooms: List[str]             # PreferredRooms (SOFT)


@dataclass
class Assignment:
    meeting_name: str
    room_name: str
    day: date
    start_dt: datetime
    end_dt: datetime


# -----------------------------
# Parsing helpers
# -----------------------------
def parse_excel_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    raise ValueError(f"Unrecognized date value: {value!r}")


def parse_excel_time(value) -> Optional[time]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, (int, float)):
        total_seconds = int(round(float(value) * 24 * 3600))
        hh = total_seconds // 3600
        mm = (total_seconds % 3600) // 60
        return time(hh % 24, mm)
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%H:%M").time()
    raise ValueError(f"Unrecognized time value: {value!r}")


def dt_combine(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute)


def overlaps(a_start: datetime, a_end: datetime, b_start: datetime, b_end: datetime) -> bool:
    return a_start < b_end and b_start < a_end


def overlap_minutes(a_start: datetime, a_end: datetime, b_start: datetime, b_end: datetime) -> int:
    if not overlaps(a_start, a_end, b_start, b_end):
        return 0
    latest_start = max(a_start, b_start)
    earliest_end = min(a_end, b_end)
    return int((earliest_end - latest_start).total_seconds() // 60)


def snap_to_grid(dt_val: datetime) -> datetime:
    minute_mod = dt_val.minute % SLOT_MINUTES
    if minute_mod != 0:
        dt_val = (dt_val + timedelta(minutes=(SLOT_MINUTES - minute_mod))).replace(second=0, microsecond=0)
    return dt_val.replace(second=0, microsecond=0)


# -----------------------------
# Load workbook inputs
# -----------------------------
def load_rooms_and_openblocks(wb) -> Tuple[Dict[str, Room], List[OpenBlock]]:
    if "Rooms" not in wb.sheetnames:
        raise ValueError("Workbook must contain a sheet named 'Rooms'.")

    ws = wb["Rooms"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header) if name is not None}

    required = ["RoomName", "Size", "Date", "OpenStart", "OpenEnd"]
    for r in required:
        if r not in idx:
            raise ValueError(f"'Rooms' sheet missing required column: {r}")

    rooms: Dict[str, Room] = {}
    blocks: List[OpenBlock] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        room_name = (row[idx["RoomName"]] or "")
        room_name = room_name.strip() if isinstance(room_name, str) else str(room_name).strip()
        if not room_name:
            continue

        size = (row[idx["Size"]] or "").strip().lower()
        if size not in SIZE_ORDER:
            raise ValueError(f"Invalid room size {size!r} for room {room_name!r}")

        day = parse_excel_date(row[idx["Date"]])
        open_start = parse_excel_time(row[idx["OpenStart"]])
        open_end = parse_excel_time(row[idx["OpenEnd"]])

        if day is None or open_start is None or open_end is None:
            raise ValueError(f"Rooms row missing Date/OpenStart/OpenEnd for room {room_name!r}")

        rooms[room_name] = Room(room_name, size)
        blocks.append(OpenBlock(room_name, day, open_start, open_end))

    return rooms, blocks


def load_meetings(wb) -> List[Meeting]:
    if "Meetings" not in wb.sheetnames:
        raise ValueError("Workbook must contain a sheet named 'Meetings'.")

    ws = wb["Meetings"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header) if name is not None}

    required = ["MeetingName", "DurationMin", "MinRoomSize"]
    for r in required:
        if r not in idx:
            raise ValueError(f"'Meetings' sheet missing required column: {r}")

    def get(col: str, row):
        return row[idx[col]] if col in idx else None

    meetings: List[Meeting] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (get("MeetingName", row) or "").strip()
        if not name:
            continue

        duration = int(get("DurationMin", row))
        min_size = (get("MinRoomSize", row) or "").strip().lower()
        if min_size not in SIZE_ORDER:
            raise ValueError(f"Invalid MinRoomSize {min_size!r} for meeting {name!r}")

        priority = int(get("Priority", row) or 0)

        preferred_day_val = get("PreferredDate", row)
        preferred_day = parse_excel_date(preferred_day_val) if preferred_day_val not in (None, "") else None

        earliest = parse_excel_time(get("EarliestStart", row))
        latest_end = parse_excel_time(get("LatestEnd", row))

        avoid_raw = (get("AvoidOverlapWith", row) or "").strip()
        avoid = [x.strip() for x in avoid_raw.split(",") if x.strip()]

        pref_rooms_raw = (get("PreferredRooms", row) or "").strip()
        preferred_rooms = [x.strip() for x in pref_rooms_raw.split(",") if x.strip()]

        meetings.append(
            Meeting(
                name=name,
                duration_min=duration,
                min_size=min_size,
                priority=priority,
                preferred_day=preferred_day,
                earliest=earliest,
                latest_end=latest_end,
                avoid_overlap_with=avoid,
                preferred_rooms=preferred_rooms,
            )
        )

    return meetings


# -----------------------------
# Availability map
# -----------------------------
def build_open_map(blocks: List[OpenBlock]) -> Dict[Tuple[str, date], List[Tuple[datetime, datetime]]]:
    open_map: Dict[Tuple[str, date], List[Tuple[datetime, datetime]]] = {}
    for b in blocks:
        open_map.setdefault((b.room_name, b.day), []).append((dt_combine(b.day, b.start), dt_combine(b.day, b.end)))

    # Merge overlaps
    for key, intervals in list(open_map.items()):
        intervals.sort()
        merged: List[List[datetime]] = []
        for s, e in intervals:
            if not merged or s > merged[-1][1]:
                merged.append([s, e])
            else:
                merged[-1][1] = max(merged[-1][1], e)
        open_map[key] = [(a, b) for a, b in merged]

    return open_map


# -----------------------------
# Candidate generation
# -----------------------------
def iter_candidate_starts_any(open_intervals: List[Tuple[datetime, datetime]], duration: timedelta) -> List[datetime]:
    starts: List[datetime] = []
    for o_start, o_end in open_intervals:
        s = snap_to_grid(o_start)
        while s + duration <= o_end:
            starts.append(s)
            s += timedelta(minutes=SLOT_MINUTES)
    return starts


# -----------------------------
# Scoring helpers
# -----------------------------
def time_window_penalty(m: Meeting, day: date, s: datetime, e: datetime) -> Tuple[int, Optional[str]]:
    """Penalty for missing preferred time window (EarliestStart/LatestEnd)."""
    if not m.earliest and not m.latest_end:
        return 0, None

    pref_start = dt_combine(day, m.earliest) if m.earliest else None
    pref_end = dt_combine(day, m.latest_end) if m.latest_end else None

    minutes_out = 0
    if pref_start and s < pref_start:
        minutes_out += int((pref_start - s).total_seconds() // 60)
    if pref_end and e > pref_end:
        minutes_out += int((e - pref_end).total_seconds() // 60)

    if minutes_out <= 0:
        return 0, None

    window = f"{m.earliest.strftime('%H:%M') if m.earliest else '—'}–{m.latest_end.strftime('%H:%M') if m.latest_end else '—'}"
    got = f"{s.strftime('%H:%M')}–{e.strftime('%H:%M')}"
    note = f"Time window (preferred: {window}, got: {got})"
    return W_TIME + minutes_out * W_TIME_PER_MIN, note


def date_penalty(m: Meeting, day: date) -> Tuple[int, Optional[str]]:
    if not m.preferred_day:
        return 0, None
    if day == m.preferred_day:
        return 0, None
    return W_DATE, f"Date (preferred: {m.preferred_day.isoformat()}, got: {day.isoformat()})"


def room_penalty(m: Meeting, room_name: str) -> Tuple[int, Optional[str]]:
    if not m.preferred_rooms:
        return 0, None
    if room_name in m.preferred_rooms:
        return 0, None
    return W_ROOM, f"Room (preferred: {', '.join(m.preferred_rooms)}, got: {room_name})"


def avoid_overlap_penalty(
    m: Meeting,
    s: datetime,
    e: datetime,
    scheduled_times: Dict[str, Tuple[datetime, datetime]],
) -> Tuple[int, List[str]]:
    """
    Soft preference: avoid overlapping listed meetings (if those meetings are already scheduled).
    Returns penalty and a list of notes for each violated overlap.
    """
    total_min = 0
    notes: List[str] = []
    for other in m.avoid_overlap_with:
        if other not in scheduled_times:
            continue
        os, oe = scheduled_times[other]
        mins = overlap_minutes(s, e, os, oe)
        if mins > 0:
            total_min += mins
            notes.append(f"Avoid overlap (overlaps '{other}' by {mins} min)")
    if total_min == 0:
        return 0, []
    return W_OVERLAP + total_min * W_OVERLAP_PER_MIN, notes


# -----------------------------
# Scheduler (greedy) with scoring + turnover buffer
# -----------------------------
def schedule_meetings(
    rooms: Dict[str, Room],
    open_map: Dict[Tuple[str, date], List[Tuple[datetime, datetime]]],
    meetings: List[Meeting],
) -> Tuple[List[Assignment], List[str]]:
    all_days: Set[date] = set(d for _, d in open_map.keys())
    if not all_days:
        return [], ["No open room blocks found."]

    # Greedy order: higher priority first, then longer, then bigger
    meetings_sorted = sorted(
        meetings,
        key=lambda m: (m.priority, m.duration_min, SIZE_ORDER[m.min_size]),
        reverse=True,
    )

    buffer_td = timedelta(minutes=ROOM_TURNOVER_MINUTES)

    # bookings per room/day
    room_bookings: Dict[Tuple[str, date], List[Tuple[datetime, datetime, str]]] = {}
    # scheduled time per meeting (for overlap preference scoring)
    meeting_time: Dict[str, Tuple[datetime, datetime]] = {}

    assignments: List[Assignment] = []
    failures: List[str] = []

    def room_fits(room: Room, min_size: str) -> bool:
        return SIZE_ORDER[room.size] >= SIZE_ORDER[min_size]

    def is_room_free_with_turnover(room_name: str, day: date, s: datetime, e: datetime) -> bool:
        # HARD: turnover buffer between meetings in same room
        for bs, be, _ in room_bookings.get((room_name, day), []):
            if overlaps(s, e, bs - buffer_td, be + buffer_td):
                return False
        return True

    def candidate_days(m: Meeting) -> List[date]:
        # Keep preferred-day-first ordering to reduce search/aid scoring ties
        days_sorted = sorted(all_days)
        if m.preferred_day and m.preferred_day in all_days:
            return [m.preferred_day] + [d for d in days_sorted if d != m.preferred_day]
        return days_sorted

    def candidate_rooms(m: Meeting) -> List[str]:
        # Keep preferred-rooms-first ordering to reduce search/aid scoring ties
        preferred = [r for r in m.preferred_rooms if r in rooms]
        remaining = [r for r in rooms.keys() if r not in preferred]
        return preferred + remaining

    def best_slot_for(m: Meeting) -> Optional[Assignment]:
        duration = timedelta(minutes=m.duration_min)

        best: Optional[Assignment] = None
        best_score: Optional[int] = None

        for day in candidate_days(m):
            for room_name in candidate_rooms(m):
                room = rooms.get(room_name)
                if not room or not room_fits(room, m.min_size):
                    continue

                intervals = open_map.get((room_name, day), [])
                if not intervals:
                    continue

                for s in iter_candidate_starts_any(intervals, duration):
                    e = s + duration

                    # HARD room conflict with turnover
                    if not is_room_free_with_turnover(room_name, day, s, e):
                        continue

                    # Compute preference score
                    score = 0

                    p_overlap, _notes_overlap = avoid_overlap_penalty(m, s, e, meeting_time)
                    score += p_overlap

                    p_time, _note_time = time_window_penalty(m, day, s, e)
                    score += p_time

                    p_date, _note_date = date_penalty(m, day)
                    score += p_date

                    p_room, _note_room = room_penalty(m, room_name)
                    score += p_room

                    # Tie-breakers (small nudges):
                    # prefer earlier times slightly, and prefer smaller (but sufficient) rooms slightly
                    score += int((s.hour * 60 + s.minute) / 15)  # earlier is better
                    score += (SIZE_ORDER[room.size] - SIZE_ORDER[m.min_size]) * 5

                    if best_score is None or score < best_score:
                        best_score = score
                        best = Assignment(m.name, room_name, day, s, e)

        return best

    for m in meetings_sorted:
        a = best_slot_for(m)
        if a is None:
            failures.append(f"Could not schedule meeting: {m.name}")
            continue

        # Commit booking
        room_bookings.setdefault((a.room_name, a.day), []).append((a.start_dt, a.end_dt, m.name))
        meeting_time[m.name] = (a.start_dt, a.end_dt)
        assignments.append(a)

    return assignments, failures


# -----------------------------
# Schedule sheet (polished)
# -----------------------------
def meeting_fill(meeting_name: str) -> PatternFill:
    palette = ["FFF2CC", "DDEBF7", "E2EFDA", "FCE4D6", "EAD1DC", "D9E1F2", "D0E0E3"]
    idx = abs(hash(meeting_name)) % len(palette)
    return PatternFill("solid", fgColor=palette[idx])


def write_schedule_sheet_polished(
    wb,
    rooms: Dict[str, Room],
    open_map: Dict[Tuple[str, date], List[Tuple[datetime, datetime]]],
    assignments: List[Assignment],
    failures: List[str],
):
    if "Schedule" in wb.sheetnames:
        del wb["Schedule"]
    ws = wb.create_sheet("Schedule")

    thin = Side(style="thin", color="9E9E9E")
    thick = Side(style="medium", color="404040")
    heavy = Side(style="thick", color="202020")

    header_fill = PatternFill("solid", fgColor="EFEFEF")
    closed_fill = PatternFill(fill_type="darkTrellis", fgColor="A6A6A6", bgColor="FFFFFF")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    by_day_room: Dict[Tuple[date, str], List[Assignment]] = {}
    days: Set[date] = set()
    for a in assignments:
        days.add(a.day)
        by_day_room.setdefault((a.day, a.room_name), []).append(a)
    days |= set(d for _, d in open_map.keys())

    days_sorted = sorted(days)
    room_names = sorted(rooms.keys())

    # time bounds from open blocks
    min_time = time(23, 59)
    max_time = time(0, 0)
    for (_, _d), intervals in open_map.items():
        for sdt, edt in intervals:
            min_time = min(min_time, sdt.time())
            max_time = max(max_time, edt.time())
    if min_time >= max_time:
        min_time, max_time = time(8, 0), time(18, 0)

    def time_slots_for_day(d: date) -> List[datetime]:
        start = dt_combine(d, min_time)
        end = dt_combine(d, max_time)
        cur = start
        out: List[datetime] = []
        while cur < end:
            out.append(cur)
            cur += timedelta(minutes=SLOT_MINUTES)
        return out

    def is_open(room_name: str, d: date, slot_start: datetime) -> bool:
        for s, e in open_map.get((room_name, d), []):
            if s <= slot_start < e:
                return True
        return False

    def room_col_border(top=False, bottom=False) -> Border:
        return Border(
            left=heavy, right=heavy,
            top=(heavy if top else thin),
            bottom=(heavy if bottom else thin),
        )

    # Header / legend
    row = 1
    ws.cell(row=row, column=1, value="Scheduling Results").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f"Scheduled: {len(assignments)}").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"Unscheduled: {len(failures)}").font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value="Legend:").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Closed time").alignment = left
    lc = ws.cell(row=row, column=3, value="")
    lc.fill = closed_fill
    lc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row += 2

    first_grid_row = row

    for d in days_sorted:
        ws.cell(row=row, column=1, value=f"Date: {d.isoformat()}").font = Font(bold=True, size=12)
        row += 1

        # header row
        tcell = ws.cell(row=row, column=1, value="Time")
        tcell.font = Font(bold=True)
        tcell.fill = header_fill
        tcell.alignment = center
        tcell.border = Border(left=thick, right=thick, top=thick, bottom=thick)

        for c, rname in enumerate(room_names, start=2):
            cell = ws.cell(row=row, column=c, value=f"{rname}\n({rooms[rname].size})")
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = room_col_border(top=True, bottom=True)
        row += 1

        slots = time_slots_for_day(d)
        slot_row_map: Dict[datetime, int] = {}

        # grid
        for slot_dt in slots:
            slot_row_map[slot_dt] = row
            is_hour = (slot_dt.minute == 0)

            tc = ws.cell(row=row, column=1, value=slot_dt.strftime("%H:%M"))
            tc.alignment = center
            tc.border = Border(left=thick, right=thick, top=(thick if is_hour else thin), bottom=thin)

            for c, rname in enumerate(room_names, start=2):
                cell = ws.cell(row=row, column=c, value="")
                cell.alignment = left
                cell.border = Border(left=heavy, right=heavy, top=(thick if is_hour else thin), bottom=thin)
                if not is_open(rname, d, slot_dt):
                    cell.fill = closed_fill
            row += 1

        # place meetings (merged blocks)
        for rname in room_names:
            col = room_names.index(rname) + 2
            room_assignments = sorted(by_day_room.get((d, rname), []), key=lambda a: a.start_dt)

            for a in room_assignments:
                start = a.start_dt
                end = a.end_dt

                if start not in slot_row_map:
                    continue
                top_row = slot_row_map[start]

                last_slot = end - timedelta(minutes=SLOT_MINUTES)
                if last_slot not in slot_row_map:
                    continue
                bottom_row = slot_row_map[last_slot]

                fill = meeting_fill(a.meeting_name)

                if bottom_row > top_row:
                    ws.merge_cells(start_row=top_row, start_column=col, end_row=bottom_row, end_column=col)

                cell = ws.cell(row=top_row, column=col)
                cell.value = f"{a.meeting_name}\n{a.start_dt.strftime('%H:%M')}–{a.end_dt.strftime('%H:%M')}"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = fill

                for rr in range(top_row, bottom_row + 1):
                    ccell = ws.cell(row=rr, column=col)
                    ccell.fill = fill
                    ccell.border = Border(
                        left=heavy, right=heavy,
                        top=(thick if rr == top_row else thin),
                        bottom=(thick if rr == bottom_row else thin),
                    )

        row += 2  # spacer between days

    # widths + freeze
    ws.column_dimensions["A"].width = 9
    for c in range(2, len(room_names) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 22

    ws.freeze_panes = ws.cell(row=first_grid_row + 1, column=2).coordinate


# -----------------------------
# Diagnostics for UNSCHEDULED (only HARD blockers)
# -----------------------------
def diagnose_unscheduled_reason(
    m: Meeting,
    rooms: Dict[str, Room],
    open_map: Dict[Tuple[str, date], List[Tuple[datetime, datetime]]],
    assignments: List[Assignment],
    all_days: Set[date],
) -> str:
    buffer_td = timedelta(minutes=ROOM_TURNOVER_MINUTES)
    duration = timedelta(minutes=m.duration_min)

    room_bookings: Dict[Tuple[str, date], List[Tuple[datetime, datetime, str]]] = {}
    for a in assignments:
        room_bookings.setdefault((a.room_name, a.day), []).append((a.start_dt, a.end_dt, a.meeting_name))

    eligible_rooms = [rn for rn, r in rooms.items() if SIZE_ORDER[r.size] >= SIZE_ORDER[m.min_size]]
    if not eligible_rooms:
        return f"No rooms meet MinRoomSize='{m.min_size}'."

    days_sorted = sorted(all_days)
    candidate_days = (
        [m.preferred_day] + [d for d in days_sorted if d != m.preferred_day]
        if (m.preferred_day and m.preferred_day in all_days)
        else days_sorted
    )

    if not any(open_map.get((rn, d)) for d in candidate_days for rn in eligible_rooms):
        return "No open blocks for any eligible room on any available date."

    def is_room_free(room_name: str, day: date, s: datetime, e: datetime) -> bool:
        for bs, be, _ in room_bookings.get((room_name, day), []):
            if overlaps(s, e, bs - buffer_td, be + buffer_td):
                return False
        return True

    total_candidates = 0
    blocked_by_turnover = 0
    blocked_by_fragmentation = 0

    for d in candidate_days:
        for rn in eligible_rooms:
            intervals = open_map.get((rn, d), [])
            if not intervals:
                continue
            starts = iter_candidate_starts_any(intervals, duration)
            if not starts:
                blocked_by_fragmentation += 1
                continue
            for s in starts:
                total_candidates += 1
                e = s + duration
                if not is_room_free(rn, d, s, e):
                    blocked_by_turnover += 1

    if total_candidates == 0:
        return "Open blocks exist, but none can fit the meeting duration on a 15-min grid."
    if blocked_by_turnover == total_candidates:
        return "All possible times conflict with existing bookings (including 15-min turnover buffer)."
    return "No feasible slot found due to hard constraints (unexpected mix of blockers)."


# -----------------------------
# Meetings sheet annotations (explicit missed preferences)
# -----------------------------
def annotate_meetings_sheet(
    wb,
    rooms: Dict[str, Room],
    open_map: Dict[Tuple[str, date], List[Tuple[datetime, datetime]]],
    assignments: List[Assignment],
    meetings: List[Meeting],
):
    ws = wb["Meetings"]

    assignment_by_name: Dict[str, Assignment] = {a.meeting_name: a for a in assignments}
    meeting_by_name: Dict[str, Meeting] = {m.name: m for m in meetings}
    all_days: Set[date] = set(d for _, d in open_map.keys())

    # scheduled times for overlap notes
    scheduled_times: Dict[str, Tuple[datetime, datetime]] = {
        a.meeting_name: (a.start_dt, a.end_dt) for a in assignments
    }

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i + 1 for i, name in enumerate(header) if name}

    if "MeetingName" not in idx:
        raise ValueError("'Meetings' sheet missing MeetingName column.")

    if "Status" not in idx:
        ws.cell(row=1, column=ws.max_column + 1, value="Status")
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i + 1 for i, name in enumerate(header) if name}

    status_col = idx["Status"]

    if "UnscheduledReason" not in idx:
        ws.insert_cols(status_col + 1)
        ws.cell(row=1, column=status_col + 1, value="UnscheduledReason")
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i + 1 for i, name in enumerate(header) if name}

    for col_name in ["AssignedRoom", "AssignedStart", "AssignedEnd", "Notes"]:
        if col_name not in idx:
            ws.cell(row=1, column=ws.max_column + 1, value=col_name)
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {name: i + 1 for i, name in enumerate(header) if name}

    red_fill = PatternFill("solid", fgColor="F8CBAD")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="C6E0B4")
    bold = Font(bold=True)

    for col_name in ["Status", "UnscheduledReason", "AssignedRoom", "AssignedStart", "AssignedEnd", "Notes"]:
        ws.cell(row=1, column=idx[col_name]).font = bold

    for r in range(2, ws.max_row + 1):
        raw = ws.cell(row=r, column=idx["MeetingName"]).value
        name = (raw or "").strip() if isinstance(raw, str) else (str(raw).strip() if raw is not None else "")
        if not name:
            continue

        m = meeting_by_name.get(name)
        a = assignment_by_name.get(name)

        status_cell = ws.cell(row=r, column=idx["Status"])
        reason_cell = ws.cell(row=r, column=idx["UnscheduledReason"])
        room_cell = ws.cell(row=r, column=idx["AssignedRoom"])
        start_cell = ws.cell(row=r, column=idx["AssignedStart"])
        end_cell = ws.cell(row=r, column=idx["AssignedEnd"])
        notes_cell = ws.cell(row=r, column=idx["Notes"])

        for c in [status_cell, reason_cell, room_cell, start_cell, end_cell, notes_cell]:
            c.value = None
            c.fill = PatternFill()

        if a is None:
            status_cell.value = "UNSCHEDULED"
            reason_cell.value = diagnose_unscheduled_reason(m, rooms, open_map, assignments, all_days) if m else "Unknown meeting definition."
            notes_cell.value = "No feasible slot found."
            for c in [status_cell, reason_cell, notes_cell]:
                c.fill = red_fill
            continue

        room_cell.value = a.room_name
        start_cell.value = a.start_dt.strftime("%Y-%m-%d %H:%M")
        end_cell.value = a.end_dt.strftime("%Y-%m-%d %H:%M")
        reason_cell.value = ""

        missed: List[str] = []

        # Preferences in importance order (for notes readability)
        if m:
            # 1) Avoid overlap (soft)
            p_ov, ov_notes = avoid_overlap_penalty(m, a.start_dt, a.end_dt, scheduled_times)
            # avoid_overlap_penalty includes overlaps with already scheduled meetings, but also includes self if present
            # (it won't, because self not in scheduled_times overlap is zero)
            missed.extend(ov_notes)

            # 2) Time window (soft)
            _, note_time = time_window_penalty(m, a.day, a.start_dt, a.end_dt)
            if note_time:
                missed.append(note_time)

            # 3) Date (soft)
            _, note_date = date_penalty(m, a.day)
            if note_date:
                missed.append(note_date)

            # 4) Room (soft)
            _, note_room = room_penalty(m, a.room_name)
            if note_room:
                missed.append(note_room)

        if missed:
            status_cell.value = "Scheduled (missed preferences)"
            status_cell.fill = yellow_fill
            notes_cell.value = "; ".join(missed)
            notes_cell.fill = yellow_fill
        else:
            status_cell.value = "Scheduled"
            status_cell.fill = green_fill
            notes_cell.value = ""

    ws.column_dimensions[get_column_letter(idx["Status"])].width = 28
    ws.column_dimensions[get_column_letter(idx["UnscheduledReason"])].width = 60
    ws.column_dimensions[get_column_letter(idx["AssignedRoom"])].width = 20
    ws.column_dimensions[get_column_letter(idx["AssignedStart"])].width = 20
    ws.column_dimensions[get_column_letter(idx["AssignedEnd"])].width = 20
    ws.column_dimensions[get_column_letter(idx["Notes"])].width = 90


# -----------------------------
# Main
# -----------------------------
def main():
    ap = argparse.ArgumentParser(description="Excel room scheduler (polished schedule + preference scoring + turnover buffer).")
    ap.add_argument("input_xlsx", help="Path to input Excel file")
    ap.add_argument("--output", default=None, help="Path to output Excel file (default: overwrite input)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input_xlsx)

    rooms, blocks = load_rooms_and_openblocks(wb)
    meetings = load_meetings(wb)
    open_map = build_open_map(blocks)

    assignments, failures = schedule_meetings(rooms, open_map, meetings)

    write_schedule_sheet_polished(wb, rooms, open_map, assignments, failures)
    annotate_meetings_sheet(wb, rooms, open_map, assignments, meetings)

    out_path = args.output or args.input_xlsx
    wb.save(out_path)

    print(f"Wrote output to: {out_path}")
    print(f"Scheduled: {len(assignments)} | Unscheduled: {len(failures)}")
    if failures:
        print("Unscheduled:")
        for f in failures:
            print(" -", f)


if __name__ == "__main__":
    main()